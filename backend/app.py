from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
import jwt
import os
import json
import io
import csv
from datetime import datetime, timedelta
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import text
from werkzeug.exceptions import BadRequest
import traceback

app = Flask(__name__)
CORS(app, origins=os.environ.get("CORS_ORIGIN", "*"))

# Config
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-change-in-production')

def _normalize_database_url(url: str) -> str:
    if not url:
        return 'sqlite:///profiles.db'
    u = url.strip()
    # Render/Heroku style URLs may start with postgres:// which SQLAlchemy may not accept.
    if u.startswith('postgres://'):
        u = 'postgresql://' + u[len('postgres://'):]
    # Ensure SSL for managed Postgres unless explicitly provided.
    if u.startswith('postgresql://') and 'sslmode=' not in u:
        u = u + ('&' if '?' in u else '?') + 'sslmode=require'
    return u

app.config['SQLALCHEMY_DATABASE_URI'] = _normalize_database_url(os.environ.get('DATABASE_URL', ''))
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 2 * 1024 * 1024  # 2MB max upload

def _to_int_or_none(val):
    if val is None:
        return None
    if isinstance(val, int):
        return val
    s = str(val).strip()
    if s == '':
        return None
    try:
        return int(s)
    except Exception:
        return None

db = SQLAlchemy(app)
bcrypt = Bcrypt(app)

# ─── Models ─────────────────────────────────────────────────────────────────

class AdminUser(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)

class Profile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    hm_id = db.Column(db.String(50), unique=True, nullable=False)
    name = db.Column(db.String(200), nullable=False)
    competency = db.Column(db.String(20))
    joining_date = db.Column(db.String(20))
    total_exp_years = db.Column(db.Integer)
    total_exp_months = db.Column(db.Integer)
    relevant_exp_years = db.Column(db.Integer)
    relevant_exp_months = db.Column(db.Integer)
    reporting_location_type = db.Column(db.String(30))
    customer_name = db.Column(db.String(200))
    customer_address = db.Column(db.Text)
    office_city = db.Column(db.String(100))
    industries = db.Column(db.Text)  # JSON array
    primary_role = db.Column(db.String(100))
    profile_pic = db.Column(db.Text)  # base64
    education = db.Column(db.Text)   # JSON array
    skills = db.Column(db.Text)      # JSON array
    certifications = db.Column(db.Text)  # JSON array
    projects = db.Column(db.Text)    # JSON array
    approved = db.Column(db.Boolean, default=False)
    approved_at = db.Column(db.DateTime)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# ─── Skills master helpers ───────────────────────────────────────────────────

def _skill_master_lookup_by_name(name: str):
    if not name:
        return None
    n = str(name).strip().lower()
    if not n:
        return None
    for s in SKILLS_MASTER:
        if s.get('skill_name', '').strip().lower() == n:
            return s
    return None

def normalize_skills_list(skills):
    """Normalize incoming/outgoing skills to structured objects.

    Backward compatible with older entries like:
    - "Python"
    - {"skill_name": "Python", ...}
    """
    if not isinstance(skills, list):
        return []

    normalized = []
    seen = set()

    for item in skills:
        if isinstance(item, str):
            obj = {
                'skill_id': None,
                'skill_name': item,
                'platform_group': None,
                'primary_secondary': 'Primary',
                'years_exp': '',
                'self_assessment': '',
            }
        elif isinstance(item, dict):
            obj = {
                'skill_id': item.get('skill_id'),
                'skill_name': item.get('skill_name') or item.get('name') or '',
                'platform_group': item.get('platform_group'),
                'primary_secondary': item.get('primary_secondary') or item.get('primary') or 'Primary',
                'years_exp': item.get('years_exp', ''),
                'self_assessment': item.get('self_assessment', ''),
            }
        else:
            continue

        # Enrich from master list if missing
        if (not obj.get('skill_id') or not obj.get('platform_group')) and obj.get('skill_name'):
            m = _skill_master_lookup_by_name(obj.get('skill_name'))
            if m:
                obj['skill_id'] = obj.get('skill_id') or m.get('skill_id')
                obj['platform_group'] = obj.get('platform_group') or m.get('platform_group')

        # Deduplicate
        dedupe_key = (obj.get('skill_id') or obj.get('skill_name') or '').strip().lower()
        if not dedupe_key or dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        normalized.append(obj)

    return normalized

# ─── Auth ────────────────────────────────────────────────────────────────────

def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        token = request.headers.get('Authorization', '').replace('Bearer ', '')
        if not token:
            return jsonify({'error': 'Token missing'}), 401
        try:
            jwt.decode(token, app.config['SECRET_KEY'], algorithms=['HS256'])
        except jwt.ExpiredSignatureError:
            return jsonify({'error': 'Token expired'}), 401
        except jwt.InvalidTokenError:
            return jsonify({'error': 'Invalid token'}), 401
        return f(*args, **kwargs)
    return decorated

@app.route('/', methods=['GET'])
def root():
    return jsonify({
        'status': 'ok',
        'service': 'Happiest Minds COE Skill Profiling Tool API'
    })

@app.route('/health', methods=['GET'])
def health():
    return jsonify({'status': 'ok'})

@app.errorhandler(BadRequest)
def handle_bad_request(e):
    if request.path.startswith('/api'):
        return jsonify({'error': 'Bad request'}), 400
    return e

@app.errorhandler(Exception)
def handle_exception(e):
    if request.path.startswith('/api'):
        print('❌ API error:', repr(e))
        traceback.print_exc()
        return jsonify({'error': 'Internal server error'}), 500
    raise e

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    data = request.json
    user = AdminUser.query.filter_by(username=data.get('username')).first()
    if user and bcrypt.check_password_hash(user.password_hash, data.get('password', '')):
        token = jwt.encode({
            'sub': user.username,
            'exp': datetime.utcnow() + timedelta(hours=8)
        }, app.config['SECRET_KEY'], algorithm='HS256')
        return jsonify({'token': token})
    return jsonify({'error': 'Invalid credentials'}), 401

# ─── Profile Routes ──────────────────────────────────────────────────────────

@app.route('/api/profile', methods=['POST'])
def submit_profile():
    data = request.get_json(silent=True)
    if not isinstance(data, dict):
        return jsonify({'error': 'Invalid JSON body'}), 400
    hm_id = data.get('hm_id', '').strip()
    if not hm_id:
        return jsonify({'error': 'Happiest Minds ID is required'}), 400

    existing = Profile.query.filter_by(hm_id=hm_id).first()
    if existing:
        # Update
        int_fields = {'total_exp_years','total_exp_months','relevant_exp_years','relevant_exp_months'}
        for field in ['name','competency','joining_date','total_exp_years','total_exp_months',
                      'relevant_exp_years','relevant_exp_months','reporting_location_type',
                      'customer_name','customer_address','office_city','primary_role','profile_pic']:
            if field in data:
                if field in int_fields:
                    setattr(existing, field, _to_int_or_none(data[field]))
                else:
                    setattr(existing, field, data[field])
        for json_field in ['industries','education','skills','certifications','projects']:
            if json_field in data:
                if json_field == 'skills':
                    setattr(existing, json_field, json.dumps(normalize_skills_list(data[json_field])))
                else:
                    setattr(existing, json_field, json.dumps(data[json_field]))
        existing.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'message': 'Profile updated successfully', 'id': existing.id})

    profile = Profile(
        hm_id=hm_id,
        name=data.get('name', ''),
        competency=data.get('competency'),
        joining_date=data.get('joining_date'),
        total_exp_years=_to_int_or_none(data.get('total_exp_years')),
        total_exp_months=_to_int_or_none(data.get('total_exp_months')),
        relevant_exp_years=_to_int_or_none(data.get('relevant_exp_years')),
        relevant_exp_months=_to_int_or_none(data.get('relevant_exp_months')),
        reporting_location_type=data.get('reporting_location_type'),
        customer_name=data.get('customer_name'),
        customer_address=data.get('customer_address'),
        office_city=data.get('office_city'),
        industries=json.dumps(data.get('industries', [])),
        primary_role=data.get('primary_role'),
        profile_pic=data.get('profile_pic'),
        education=json.dumps(data.get('education', [])),
        skills=json.dumps(normalize_skills_list(data.get('skills', []))),
        certifications=json.dumps(data.get('certifications', [])),
        projects=json.dumps(data.get('projects', [])),
    )
    db.session.add(profile)
    db.session.commit()
    return jsonify({'message': 'Profile submitted successfully', 'id': profile.id}), 201

@app.route('/api/profile/<hm_id>', methods=['GET'])
def get_profile(hm_id):
    profile = Profile.query.filter_by(hm_id=hm_id).first()
    if not profile:
        return jsonify({'error': 'Profile not found'}), 404
    if not profile.approved:
        return jsonify({'error': 'Profile pending admin approval'}), 403
    return jsonify(profile_to_dict(profile))

# ─── Admin Routes ────────────────────────────────────────────────────────────

@app.route('/api/admin/profiles', methods=['GET'])
@token_required
def list_profiles():
    search = request.args.get('search', '')
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    query = Profile.query
    if search:
        query = query.filter(
            db.or_(Profile.name.ilike(f'%{search}%'), Profile.hm_id.ilike(f'%{search}%'))
        )
    paginated = query.order_by(Profile.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
    return jsonify({
        'profiles': [profile_to_dict(p) for p in paginated.items],
        'total': paginated.total,
        'pages': paginated.pages,
        'page': page
    })

@app.route('/api/admin/profiles/<int:profile_id>', methods=['DELETE'])
@token_required
def delete_profile(profile_id):
    profile = Profile.query.get_or_404(profile_id)
    db.session.delete(profile)
    db.session.commit()
    return jsonify({'message': 'Deleted'})

@app.route('/api/admin/profiles/<int:profile_id>/approval', methods=['PATCH'])
@token_required
def set_profile_approval(profile_id):
    profile = Profile.query.get_or_404(profile_id)
    data = request.json or {}
    approved = bool(data.get('approved'))
    profile.approved = approved
    profile.approved_at = datetime.utcnow() if approved else None
    db.session.commit()
    return jsonify({'message': 'Updated', 'approved': profile.approved, 'approved_at': profile.approved_at.isoformat() if profile.approved_at else None})

def _safe_json_loads(val, default):
    try:
        if val is None:
            return default
        if isinstance(val, (list, dict)):
            return val
        s = str(val).strip()
        if not s:
            return default
        return json.loads(s)
    except Exception:
        return default

def _fmt_join(values):
    parts = []
    for v in values or []:
        s = (str(v) if v is not None else '').strip()
        if s:
            parts.append(s)
    return '; '.join(parts)

def _fmt_education(education):
    rows = []
    for e in education or []:
        if not isinstance(e, dict):
            continue
        degree = (e.get('degree') or '').strip()
        spec = (e.get('specialisation') or '').strip()
        inst = (e.get('institution') or '').strip()
        year = (e.get('year') or '').strip()
        grade = (e.get('grade') or '').strip()
        line = ' | '.join([p for p in [degree, spec, inst, year, grade] if p])
        if line:
            rows.append(line)
    return '\n'.join(rows)

def _fmt_skills(skills):
    skills_norm = normalize_skills_list(skills or [])
    rows = []
    for s in skills_norm:
        if not isinstance(s, dict):
            continue
        skill_id = (s.get('skill_id') or '').strip()
        name = (s.get('skill_name') or '').strip()
        group = (s.get('platform_group') or '').strip()
        ps = (s.get('primary_secondary') or '').strip()
        years = (str(s.get('years_exp')) if s.get('years_exp') is not None else '').strip()
        sa = (s.get('self_assessment') or '').strip()
        main = ' '.join([p for p in [skill_id, name] if p]).strip()
        meta = _fmt_join([group, ps, (f"{years}y" if years else ''), sa])
        line = f"{main} — {meta}" if meta else main
        if line:
            rows.append(line)
    return '\n'.join(rows)

def _fmt_certs(certs):
    rows = []
    for c in certs or []:
        if not isinstance(c, dict):
            continue
        name = (c.get('name') or '').strip()
        provider = (c.get('provider') or '').strip()
        date = (c.get('date') or '').strip()
        expiry = (c.get('expiry') or '').strip()
        meta = _fmt_join([provider, (f"Obtained: {date}" if date else ''), (f"Expiry: {expiry}" if expiry else '')])
        line = f"{name} — {meta}" if meta else name
        if line:
            rows.append(line)
    return '\n'.join(rows)

def _fmt_projects(projects):
    rows = []
    for pr in projects or []:
        if not isinstance(pr, dict):
            continue
        title = (pr.get('title') or '').strip()
        role = (pr.get('role') or '').strip()
        duration = (pr.get('duration') or '').strip()
        tools = (pr.get('tools') or '').strip()
        desc = (pr.get('description') or '').strip()
        resp = (pr.get('responsibility') or '').strip()
        awards = (pr.get('awards') or '').strip()
        header = ' | '.join([p for p in [title, role, duration] if p]).strip()
        details = []
        if tools:
            details.append(f"Tools: {tools}")
        if desc:
            details.append(f"Desc: {desc}")
        if resp:
            details.append(f"Resp: {resp}")
        if awards:
            details.append(f"Awards: {awards}")
        if header and details:
            rows.append(header + "\n" + '\n'.join(details))
        elif header:
            rows.append(header)
        elif details:
            rows.append('\n'.join(details))
    return '\n\n'.join(rows)

@app.route('/api/admin/export/csv', methods=['GET'])
@token_required
def export_csv():
    profiles = Profile.query.all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        'HM ID','Name','Competency','Joining Date',
        'Total Exp (Y)','Total Exp (M)','Relevant Exp (Y)','Relevant Exp (M)',
        'Location Type','Customer Name','Customer Address','Office City',
        'Primary Role','Industries',
        'Education',
        'Skills (Detailed)',
        'Certifications (Detailed)',
        'Projects (Detailed)',
        'Has Profile Pic',
        'Approved','Approved At','Created At','Updated At'
    ])
    for p in profiles:
        industries = _safe_json_loads(p.industries, [])
        education = _safe_json_loads(p.education, [])
        skills = _safe_json_loads(p.skills, [])
        certs = _safe_json_loads(p.certifications, [])
        projects = _safe_json_loads(p.projects, [])

        industries_str = _fmt_join(industries)
        education_str = _fmt_education(education)
        skills_str = _fmt_skills(skills)
        certs_str = _fmt_certs(certs)
        projects_str = _fmt_projects(projects)

        writer.writerow([
            p.hm_id, p.name, p.competency, p.joining_date,
            p.total_exp_years, p.total_exp_months, p.relevant_exp_years, p.relevant_exp_months,
            p.reporting_location_type, p.customer_name, p.customer_address, p.office_city,
            p.primary_role, industries_str,
            education_str,
            skills_str,
            certs_str,
            projects_str,
            bool(p.profile_pic),
            bool(p.approved),
            p.approved_at.isoformat() if p.approved_at else '',
            p.created_at.isoformat() if p.created_at else '',
            p.updated_at.isoformat() if p.updated_at else ''
        ])
    output.seek(0)
    return send_file(io.BytesIO(output.getvalue().encode()), mimetype='text/csv',
                     as_attachment=True, download_name='profiles.csv')

@app.route('/api/admin/export/excel', methods=['GET'])
@token_required
def export_excel():
    profiles = Profile.query.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Profiles'

    headers = [
        'HM ID','Name','Competency','Joining Date',
        'Total Exp (Y)','Total Exp (M)','Relevant Exp (Y)','Relevant Exp (M)',
        'Location Type','Customer Name','Customer Address','Office City',
        'Primary Role','Industries',
        'Education',
        'Skills (Detailed)',
        'Certifications (Detailed)',
        'Projects (Detailed)',
        'Has Profile Pic',
        'Approved','Approved At','Created At','Updated At'
    ]
    header_fill = PatternFill(start_color='1F6E3C', end_color='1F6E3C', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for p in profiles:
        industries = _safe_json_loads(p.industries, [])
        education = _safe_json_loads(p.education, [])
        skills = _safe_json_loads(p.skills, [])
        certs = _safe_json_loads(p.certifications, [])
        projects = _safe_json_loads(p.projects, [])

        industries_str = _fmt_join(industries)
        education_str = _fmt_education(education)
        skills_str = _fmt_skills(skills)
        certs_str = _fmt_certs(certs)
        projects_str = _fmt_projects(projects)

        ws.append([
            p.hm_id, p.name, p.competency, p.joining_date,
            p.total_exp_years, p.total_exp_months, p.relevant_exp_years, p.relevant_exp_months,
            p.reporting_location_type, p.customer_name, p.customer_address, p.office_city,
            p.primary_role, industries_str,
            education_str,
            skills_str,
            certs_str,
            projects_str,
            bool(p.profile_pic),
            bool(p.approved),
            p.approved_at.isoformat() if p.approved_at else '',
            p.created_at.isoformat() if p.created_at else '',
            p.updated_at.isoformat() if p.updated_at else ''
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='profiles.xlsx')

@app.route('/api/admin/stats', methods=['GET'])
@token_required
def stats():
    total = Profile.query.count()
    roles = db.session.query(Profile.primary_role, db.func.count(Profile.id)).group_by(Profile.primary_role).all()
    competencies = db.session.query(Profile.competency, db.func.count(Profile.id)).group_by(Profile.competency).all()
    return jsonify({
        'total_profiles': total,
        'by_role': {r: c for r, c in roles if r},
        'by_competency': {comp: cnt for comp, cnt in competencies if comp}
    })

# ─── Skills list ─────────────────────────────────────────────────────────────

@app.route('/api/skills', methods=['GET'])
def get_skills():
    q = request.args.get('q', '').lower()
    group = request.args.get('group', '').strip().lower()
    skills = SKILLS_MASTER
    if q:
        skills = [s for s in skills if q in (s.get('skill_name', '').lower())]
    if group:
        skills = [s for s in skills if group in (s.get('platform_group', '').lower())]
    return jsonify(skills[:50])

# ─── Helpers ─────────────────────────────────────────────────────────────────

def profile_to_dict(p):
    return {
        'id': p.id, 'hm_id': p.hm_id, 'name': p.name, 'competency': p.competency,
        'joining_date': p.joining_date, 'total_exp_years': p.total_exp_years,
        'total_exp_months': p.total_exp_months, 'relevant_exp_years': p.relevant_exp_years,
        'relevant_exp_months': p.relevant_exp_months,
        'reporting_location_type': p.reporting_location_type,
        'customer_name': p.customer_name, 'customer_address': p.customer_address,
        'office_city': p.office_city, 'industries': json.loads(p.industries or '[]'),
        'primary_role': p.primary_role, 'profile_pic': p.profile_pic,
        'education': json.loads(p.education or '[]'),
        'skills': normalize_skills_list(json.loads(p.skills or '[]')),
        'certifications': json.loads(p.certifications or '[]'),
        'projects': json.loads(p.projects or '[]'),
        'approved': bool(p.approved),
        'approved_at': p.approved_at.isoformat() if p.approved_at else None,
        'created_at': p.created_at.isoformat(), 'updated_at': p.updated_at.isoformat()
    }

SKILLS_MASTER = [
    {"skill_id": "SK00001", "skill_name": "ANN", "platform_group": "AI-ML"},
    {"skill_id": "SK00002", "skill_name": "CNN", "platform_group": "AI-ML"},
    {"skill_id": "SK00003", "skill_name": "Computer Vision", "platform_group": "AI-ML"},
    {"skill_id": "SK00004", "skill_name": "Deep Learning", "platform_group": "AI-ML"},
    {"skill_id": "SK00005", "skill_name": "Keras", "platform_group": "AI-ML"},
    {"skill_id": "SK00006", "skill_name": "LangChain", "platform_group": "AI-ML"},
    {"skill_id": "SK00007", "skill_name": "LangGraph", "platform_group": "AI-ML"},
    {"skill_id": "SK00008", "skill_name": "LightGBM", "platform_group": "AI-ML"},
    {"skill_id": "SK00009", "skill_name": "MLflow", "platform_group": "AI-ML"},
    {"skill_id": "SK00010", "skill_name": "Machine Learning", "platform_group": "AI-ML"},
    {"skill_id": "SK00011", "skill_name": "Natural Language Processing", "platform_group": "AI-ML"},
    {"skill_id": "SK00012", "skill_name": "PyTorch", "platform_group": "AI-ML"},
    {"skill_id": "SK00013", "skill_name": "RAG", "platform_group": "AI-ML"},
    {"skill_id": "SK00014", "skill_name": "Scikit-learn", "platform_group": "AI-ML"},
    {"skill_id": "SK00015", "skill_name": "TensorFlow", "platform_group": "AI-ML"},
    {"skill_id": "SK00016", "skill_name": "Time Series Forecasting", "platform_group": "AI-ML"},
    {"skill_id": "SK00017", "skill_name": "XGBoost", "platform_group": "AI-ML"},
    {"skill_id": "SK00018", "skill_name": "Hugging Face Transformers", "platform_group": "AI-ML"},
    {"skill_id": "SK00019", "skill_name": "MLflow (Python)", "platform_group": "AI-ML"},
    {"skill_id": "SK00020", "skill_name": "Prompt Engineering", "platform_group": "AI-ML"},
    {"skill_id": "SK00021", "skill_name": "AWS Bedrock", "platform_group": "AWS"},
    {"skill_id": "SK00022", "skill_name": "AWS CloudFormation", "platform_group": "AWS"},
    {"skill_id": "SK00023", "skill_name": "AWS CloudWatch", "platform_group": "AWS"},
    {"skill_id": "SK00024", "skill_name": "AWS Glue", "platform_group": "AWS"},
    {"skill_id": "SK00025", "skill_name": "AWS IAM", "platform_group": "AWS"},
    {"skill_id": "SK00026", "skill_name": "AWS Lambda", "platform_group": "AWS"},
    {"skill_id": "SK00027", "skill_name": "AWS SageMaker", "platform_group": "AWS"},
    {"skill_id": "SK00028", "skill_name": "Amazon Athena", "platform_group": "AWS"},
    {"skill_id": "SK00029", "skill_name": "Amazon EC2", "platform_group": "AWS"},
    {"skill_id": "SK00030", "skill_name": "Amazon EMR", "platform_group": "AWS"},
    {"skill_id": "SK00031", "skill_name": "Amazon Redshift", "platform_group": "AWS"},
    {"skill_id": "SK00032", "skill_name": "Amazon S3", "platform_group": "AWS"},
    {"skill_id": "SK00033", "skill_name": "AWS KMS", "platform_group": "AWS"},
    {"skill_id": "SK00034", "skill_name": "Amazon Kinesis", "platform_group": "AWS"},
    {"skill_id": "SK00035", "skill_name": "Databricks (AWS)", "platform_group": "AWS"},
    {"skill_id": "SK00036", "skill_name": "Azure Blob Storage", "platform_group": "Azure"},
    {"skill_id": "SK00037", "skill_name": "Azure Data Factory", "platform_group": "Azure"},
    {"skill_id": "SK00038", "skill_name": "Azure Data Lake Storage Gen2", "platform_group": "Azure"},
    {"skill_id": "SK00039", "skill_name": "Azure Databricks", "platform_group": "Azure"},
    {"skill_id": "SK00040", "skill_name": "Azure DevOps", "platform_group": "Azure"},
    {"skill_id": "SK00041", "skill_name": "Azure Event Hub", "platform_group": "Azure"},
    {"skill_id": "SK00042", "skill_name": "Azure Functions", "platform_group": "Azure"},
    {"skill_id": "SK00043", "skill_name": "Azure Key Vault", "platform_group": "Azure"},
    {"skill_id": "SK00044", "skill_name": "Azure Machine Learning", "platform_group": "Azure"},
    {"skill_id": "SK00045", "skill_name": "Azure OpenAI", "platform_group": "Azure"},
    {"skill_id": "SK00046", "skill_name": "Azure Synapse Analytics", "platform_group": "Azure"},
    {"skill_id": "SK00047", "skill_name": "Business requirement analysis", "platform_group": "Base"},
    {"skill_id": "SK00048", "skill_name": "Data Science", "platform_group": "Base"},
    {"skill_id": "SK00049", "skill_name": "FastAPI", "platform_group": "Base"},
    {"skill_id": "SK00050", "skill_name": "Flask", "platform_group": "Base"},
    {"skill_id": "SK00051", "skill_name": "Google BigQuery", "platform_group": "Base"},
    {"skill_id": "SK00052", "skill_name": "Java", "platform_group": "Base"},
    {"skill_id": "SK00053", "skill_name": "JavaScript", "platform_group": "Base"},
    {"skill_id": "SK00054", "skill_name": "Microsoft Fabric", "platform_group": "Base"},
    {"skill_id": "SK00055", "skill_name": "PySpark", "platform_group": "Base"},
    {"skill_id": "SK00056", "skill_name": "Python", "platform_group": "Base"},
    {"skill_id": "SK00057", "skill_name": "R", "platform_group": "Base"},
    {"skill_id": "SK00058", "skill_name": "Scala", "platform_group": "Base"},
    {"skill_id": "SK00059", "skill_name": "Snowflake", "platform_group": "Base"},
    {"skill_id": "SK00060", "skill_name": "Async Programming (asyncio)", "platform_group": "Base"},
    {"skill_id": "SK00061", "skill_name": "Auto Loader", "platform_group": "Base"},
    {"skill_id": "SK00062", "skill_name": "Bash", "platform_group": "Base"},
    {"skill_id": "SK00063", "skill_name": "C++", "platform_group": "Base"},
    {"skill_id": "SK00064", "skill_name": "Cassandra", "platform_group": "Base"},
    {"skill_id": "SK00065", "skill_name": "Data Vault Modeling", "platform_group": "Base"},
    {"skill_id": "SK00066", "skill_name": "Dimensional Modeling", "platform_group": "Base"},
    {"skill_id": "SK00067", "skill_name": "Dockerizing Python Apps", "platform_group": "Base"},
    {"skill_id": "SK00068", "skill_name": "DynamoDB", "platform_group": "Base"},
    {"skill_id": "SK00069", "skill_name": "IBM DB2", "platform_group": "Base"},
    {"skill_id": "SK00070", "skill_name": "Informatica PowerCenter", "platform_group": "Base"},
    {"skill_id": "SK00071", "skill_name": "LangChain (Python)", "platform_group": "Base"},
    {"skill_id": "SK00072", "skill_name": "MongoDB", "platform_group": "Base"},
    {"skill_id": "SK00073", "skill_name": "Neo4j", "platform_group": "Base"},
    {"skill_id": "SK00074", "skill_name": "NumPy", "platform_group": "Base"},
    {"skill_id": "SK00075", "skill_name": "OOP in Python", "platform_group": "Base"},
    {"skill_id": "SK00076", "skill_name": "Oracle", "platform_group": "Base"},
    {"skill_id": "SK00077", "skill_name": "Pandas", "platform_group": "Base"},
    {"skill_id": "SK00078", "skill_name": "Pentaho PDI", "platform_group": "Base"},
    {"skill_id": "SK00079", "skill_name": "PowerShell", "platform_group": "Base"},
    {"skill_id": "SK00080", "skill_name": "PyTest", "platform_group": "Base"},
    {"skill_id": "SK00081", "skill_name": "Query Optimization", "platform_group": "Base"},
    {"skill_id": "SK00082", "skill_name": "Redis", "platform_group": "Base"},
    {"skill_id": "SK00083", "skill_name": "Stored Procedures", "platform_group": "Base"},
    {"skill_id": "SK00084", "skill_name": "Talend", "platform_group": "Base"},
    {"skill_id": "SK00085", "skill_name": "Teradata", "platform_group": "Base"},
    {"skill_id": "SK00086", "skill_name": "Window Functions", "platform_group": "Base"},
    {"skill_id": "SK00087", "skill_name": "DBT", "platform_group": "Base"},
    {"skill_id": "SK00088", "skill_name": "Power BI", "platform_group": "BI"},
    {"skill_id": "SK00089", "skill_name": "QuickSight", "platform_group": "BI"},
    {"skill_id": "SK00090", "skill_name": "SSAS", "platform_group": "BI"},
    {"skill_id": "SK00091", "skill_name": "SSIS", "platform_group": "BI"},
    {"skill_id": "SK00092", "skill_name": "SSRS", "platform_group": "BI"},
    {"skill_id": "SK00093", "skill_name": "Tableau", "platform_group": "BI"},
    {"skill_id": "SK00094", "skill_name": "DAX", "platform_group": "BI"},
    {"skill_id": "SK00095", "skill_name": "Apache NiFi", "platform_group": "BigData"},
    {"skill_id": "SK00096", "skill_name": "Hadoop", "platform_group": "BigData"},
    {"skill_id": "SK00097", "skill_name": "Hive", "platform_group": "BigData"},
    {"skill_id": "SK00098", "skill_name": "Kafka", "platform_group": "BigData"},
    {"skill_id": "SK00099", "skill_name": "Kafka Streams", "platform_group": "BigData"},
    {"skill_id": "SK00100", "skill_name": "Spark", "platform_group": "BigData"},
    {"skill_id": "SK00101", "skill_name": "Airflow", "platform_group": "BigData"},
    {"skill_id": "SK00102", "skill_name": "Apache Iceberg", "platform_group": "BigData"},
    {"skill_id": "SK00103", "skill_name": "Apache Kafka", "platform_group": "BigData"},
    {"skill_id": "SK00104", "skill_name": "Apache Spark", "platform_group": "BigData"},
    {"skill_id": "SK00105", "skill_name": "Dask", "platform_group": "BigData"},
    {"skill_id": "SK00106", "skill_name": "Structured Streaming", "platform_group": "BigData"},
    {"skill_id": "SK00107", "skill_name": "Databricks", "platform_group": "Databricks"},
    {"skill_id": "SK00108", "skill_name": "Delta Lake", "platform_group": "Databricks"},
    {"skill_id": "SK00109", "skill_name": "Delta Live Tables", "platform_group": "Databricks"},
    {"skill_id": "SK00110", "skill_name": "Unity Catalog", "platform_group": "Databricks"},
    {"skill_id": "SK00111", "skill_name": "Databricks (GCP)", "platform_group": "Databricks"},
    {"skill_id": "SK00112", "skill_name": "Databricks CLI", "platform_group": "Databricks"},
    {"skill_id": "SK00113", "skill_name": "Databricks Feature Store", "platform_group": "Databricks"},
    {"skill_id": "SK00114", "skill_name": "Databricks REST API", "platform_group": "Databricks"},
    {"skill_id": "SK00115", "skill_name": "Databricks Runtime", "platform_group": "Databricks"},
    {"skill_id": "SK00116", "skill_name": "Databricks Runtime for ML", "platform_group": "Databricks"},
    {"skill_id": "SK00117", "skill_name": "Databricks SQL", "platform_group": "Databricks"},
    {"skill_id": "SK00118", "skill_name": "Databricks SQL Warehouse", "platform_group": "Databricks"},
    {"skill_id": "SK00119", "skill_name": "Mosaic AI", "platform_group": "Databricks"},
    {"skill_id": "SK00120", "skill_name": "Docker", "platform_group": "DevOps"},
    {"skill_id": "SK00121", "skill_name": "Git", "platform_group": "DevOps"},
    {"skill_id": "SK00122", "skill_name": "Jenkins", "platform_group": "DevOps"},
    {"skill_id": "SK00123", "skill_name": "Kubernetes", "platform_group": "DevOps"},
    {"skill_id": "SK00124", "skill_name": "Terraform", "platform_group": "DevOps"},
    {"skill_id": "SK00125", "skill_name": "Autosys", "platform_group": "DevOps"},
    {"skill_id": "SK00126", "skill_name": "Cloud Composer", "platform_group": "DevOps"},
    {"skill_id": "SK00127", "skill_name": "Control-M", "platform_group": "DevOps"},
    {"skill_id": "SK00128", "skill_name": "Banking Analytics", "platform_group": "Domain"},
    {"skill_id": "SK00129", "skill_name": "Bioinformatics", "platform_group": "Domain"},
    {"skill_id": "SK00130", "skill_name": "Data Analytics", "platform_group": "Domain"},
    {"skill_id": "SK00131", "skill_name": "Gene Annotation", "platform_group": "Domain"},
    {"skill_id": "SK00132", "skill_name": "Healthcare AI", "platform_group": "Domain"},
    {"skill_id": "SK00133", "skill_name": "Human and Bacterial Genomics", "platform_group": "Domain"},
    {"skill_id": "SK00134", "skill_name": "Industrial IoT", "platform_group": "Domain"},
    {"skill_id": "SK00135", "skill_name": "Marketing Analytics", "platform_group": "Domain"},
    {"skill_id": "SK00136", "skill_name": "Metagenomics", "platform_group": "Domain"},
    {"skill_id": "SK00137", "skill_name": "Protein Structure Modelling", "platform_group": "Domain"},
    {"skill_id": "SK00138", "skill_name": "Single Cell Omics", "platform_group": "Domain"},
    {"skill_id": "SK00139", "skill_name": "Vibration Analytics", "platform_group": "Domain"},
    {"skill_id": "SK00140", "skill_name": "Business Intelligence", "platform_group": "Functional"},
    {"skill_id": "SK00141", "skill_name": "Capacity Planning", "platform_group": "Functional"},
    {"skill_id": "SK00142", "skill_name": "Data Architecture", "platform_group": "Functional"},
    {"skill_id": "SK00143", "skill_name": "Data Engineering", "platform_group": "Functional"},
    {"skill_id": "SK00144", "skill_name": "Data Governance", "platform_group": "Functional"},
    {"skill_id": "SK00145", "skill_name": "Data Migration", "platform_group": "Functional"},
    {"skill_id": "SK00146", "skill_name": "Data Modernization", "platform_group": "Functional"},
    {"skill_id": "SK00147", "skill_name": "Designing scalable architecture", "platform_group": "Functional"},
    {"skill_id": "SK00148", "skill_name": "End-to-end analytics solution designing", "platform_group": "Functional"},
    {"skill_id": "SK00149", "skill_name": "Performance Optimization", "platform_group": "Functional"},
    {"skill_id": "SK00150", "skill_name": "Root Cause Analysis", "platform_group": "Functional"},
    {"skill_id": "SK00151", "skill_name": "Agile", "platform_group": "Soft"},
    {"skill_id": "SK00152", "skill_name": "Consulting", "platform_group": "Soft"},
    {"skill_id": "SK00153", "skill_name": "Insight generation", "platform_group": "Soft"},
    {"skill_id": "SK00154", "skill_name": "Pre-Sales", "platform_group": "Soft"},
    {"skill_id": "SK00155", "skill_name": "Product Management", "platform_group": "Soft"},
    {"skill_id": "SK00156", "skill_name": "Requirement gathering", "platform_group": "Soft"},
    {"skill_id": "SK00157", "skill_name": "Research manuscript writing", "platform_group": "Soft"},
    {"skill_id": "SK00158", "skill_name": "Scrum planning", "platform_group": "Soft"},
    {"skill_id": "SK00159", "skill_name": "Sprint planning", "platform_group": "Soft"},
    {"skill_id": "SK00160", "skill_name": "Stakeholder management", "platform_group": "Soft"},
    {"skill_id": "SK00161", "skill_name": "Team management", "platform_group": "Soft"},
    {"skill_id": "SK00162", "skill_name": "PL/SQL", "platform_group": "SQL"},
    {"skill_id": "SK00163", "skill_name": "SQL", "platform_group": "SQL"},
    {"skill_id": "SK00164", "skill_name": "Spark SQL", "platform_group": "SQL"},
    {"skill_id": "SK00165", "skill_name": "T-SQL", "platform_group": "SQL"},
    {"skill_id": "SK00166", "skill_name": "ANSI SQL", "platform_group": "SQL"},
    {"skill_id": "SK00167", "skill_name": "BigQuery SQL", "platform_group": "SQL"},
    {"skill_id": "SK00168", "skill_name": "MySQL", "platform_group": "SQL"},
    {"skill_id": "SK00169", "skill_name": "PostgreSQL", "platform_group": "SQL"},
    {"skill_id": "SK00170", "skill_name": "Redshift SQL", "platform_group": "SQL"},
    {"skill_id": "SK00171", "skill_name": "SQL Server", "platform_group": "SQL"},
    {"skill_id": "SK00172", "skill_name": "Snowflake SQL", "platform_group": "SQL"},
]

# ─── Init ─────────────────────────────────────────────────────────────────────

def create_default_admin():
    with app.app_context():
        db.create_all()

        # Lightweight schema migration for existing DBs (SQLite/Postgres)
        try:
            uri = app.config.get('SQLALCHEMY_DATABASE_URI', '')
            if uri.startswith('sqlite'):
                cols = [r[1] for r in db.session.execute(text("PRAGMA table_info(profile)")).fetchall()]
                if 'approved' not in cols:
                    db.session.execute(text("ALTER TABLE profile ADD COLUMN approved BOOLEAN DEFAULT 0"))
                if 'approved_at' not in cols:
                    db.session.execute(text("ALTER TABLE profile ADD COLUMN approved_at DATETIME"))
                db.session.commit()
            elif uri.startswith('postgres'):
                cols = [r[0] for r in db.session.execute(text("""
                    SELECT column_name
                    FROM information_schema.columns
                    WHERE table_name='profile'
                """)).fetchall()]
                if 'approved' not in cols:
                    db.session.execute(text("ALTER TABLE profile ADD COLUMN approved BOOLEAN DEFAULT FALSE"))
                if 'approved_at' not in cols:
                    db.session.execute(text("ALTER TABLE profile ADD COLUMN approved_at TIMESTAMP"))
                db.session.commit()
        except Exception:
            db.session.rollback()

        if not AdminUser.query.filter_by(username='admin').first():
            admin_password = os.environ.get('ADMIN_PASSWORD', 'admin123')
            hashed = bcrypt.generate_password_hash(admin_password).decode('utf-8')
            db.session.add(AdminUser(username='admin', password_hash=hashed))
            db.session.commit()
            print(f"✅ Default admin created. Username: admin, Password: {admin_password}")

# Ensure DB/tables exist when running under gunicorn/import (Render)
if os.environ.get('SKIP_DB_INIT', '').strip().lower() not in {'1', 'true', 'yes'}:
    try:
        create_default_admin()
    except Exception:
        # Don't crash the server on init failures; requests will surface errors.
        pass

if __name__ == '__main__':
    app.run(debug=False, host='0.0.0.0', port=int(os.environ.get('PORT', 5050)))
